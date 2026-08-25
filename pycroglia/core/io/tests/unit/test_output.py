import pytest
import json
from openpyxl import load_workbook

from pycroglia.core.io.output import (
    AnalysisSummary,
    AnalysisSummaryConfig,
    CellAnalysis,
    CellAnalysisConfig,
    FullAnalysis,
    ExcelOutput,
    JSONOutput,
    BranchLengthsXlsxOutput,
    OutputWriter,
    AnalysisMetadata
)


@pytest.fixture
def sample_full_analysis():
    """Fixture that returns a sample FullAnalysis object for testing."""
    summary = AnalysisSummary(
        file="testfile.tif",
        avg_centroid_distance=12.5,
        total_territorial_volume=1000.0,
        total_unoccupied_volume=500.0,
        percent_occupied_volume=66.7,
    )
    cells = [
        CellAnalysis(
            cell_territory_volume=100.0,
            cell_volume=50.0,
            ramification_index=1.2,
            number_of_endpoints=5,
            number_of_branches=3,
            avg_branch_length=10.0,
            max_branch_length=15.0,
            min_branch_length=5.0,
            branch_analysis={
                "branch_lengths": [8.5, 12.3, 9.2],  # 3 branches
            },
        ),
        CellAnalysis(
            cell_territory_volume=110.0,
            cell_volume=55.0,
            ramification_index=1.3,
            number_of_endpoints=6,
            number_of_branches=4,
            avg_branch_length=11.0,
            max_branch_length=16.0,
            min_branch_length=6.0,
            branch_analysis={
                "branch_lengths": [10.1, 11.5, 9.8, 12.6],  # 4 branches
            },
        ),
    ]
    return FullAnalysis(summary=summary, cells=cells)


@pytest.fixture
def analysis_metadata():
    """Fixture that returns a sample AnalysisMetadata object for testing."""
    return AnalysisMetadata(
        file_path='test_path',
        gray_filter_value=2.5, 
        min_size=2,
        erosion_radius=4,
        scale=4, 
        z_scale=2,
        segmented_cell_indices=[2,6],
        selected_cell_indices=[2,7],
        rejected_cell_indices=[2,22]
    )


def test_excel_output_write(tmp_path, sample_full_analysis, analysis_metadata):
    """Test that ExcelOutput.write creates a valid Excel file with correct data."""
    file_path = tmp_path / "output_test"
    excel_writer = ExcelOutput()
    excel_writer.write(str(file_path), sample_full_analysis, analysis_metadata)

    # Check file exists
    output_file = file_path.with_suffix(".xlsx")
    assert output_file.exists()

    # Check contents
    wb = load_workbook(output_file)
    # Summary sheet
    ws_summary = wb[excel_writer.summary_title]
    summary_rows = list(ws_summary.iter_rows(values_only=True))
    assert summary_rows[0][0] == excel_writer.summary_config.file_txt
    assert summary_rows[0][1] == sample_full_analysis.summary.file
    # PerCell sheet: column 0 is "Index", config headers start at column 1
    ws_per_cell = wb[excel_writer.per_cell_title]
    per_cell_rows = list(ws_per_cell.iter_rows(values_only=True))
    assert per_cell_rows[0][0] == "Index"
    assert per_cell_rows[0][1] == excel_writer.per_cell_config.cell_territory_volume_txt
    assert per_cell_rows[1][1] == sample_full_analysis.cells[0].cell_territory_volume
    assert per_cell_rows[2][2] == sample_full_analysis.cells[1].cell_volume


def test_excel_otuput_write_custom_config(tmp_path, sample_full_analysis, analysis_metadata):
    """Test that ExcelOutput.write with custom config writes correct headers."""
    summary_config = AnalysisSummaryConfig(
        file_txt="file",
        avg_centroid_distance_txt="distance",
        total_territorial_volume_txt="territory",
        total_unoccupied_volume_txt="unoccupied",
        percent_occupied_volume_txt="percentage",
    )
    per_cell_config = CellAnalysisConfig(
        cell_territory_volume_txt="cell_territory",
        cell_volume_txt="cell_volume",
        ramification_index_txt="ramification",
        number_of_endpoints_txt="endpoints",
        number_of_branches_txt="branches",
        avg_branch_length_txt="avg_length",
        max_branch_length_txt="max_length",
        min_branch_length_txt="min_length",
    )
    file_path = tmp_path / "output_custom"
    excel_writer = ExcelOutput(
        summary_config=summary_config,
        per_cell_config=per_cell_config,
        summary_title="Summary",
        per_cell_title="PerCell",
    )
    excel_writer.write(str(file_path), sample_full_analysis, analysis_metadata)
    output_file = file_path.with_suffix(".xlsx")
    wb = load_workbook(output_file)
    ws_summary = wb["Summary"]
    summary_rows = list(ws_summary.iter_rows(values_only=True))
    assert summary_rows[0][0] == "file"
    ws_per_cell = wb["PerCell"]
    per_cell_rows = list(ws_per_cell.iter_rows(values_only=True))
    assert per_cell_rows[0][0] == "Index"
    assert per_cell_rows[0][1] == "cell_territory"

def test_json_output_write(tmp_path, sample_full_analysis, analysis_metadata):
    """Test that JSONOutput.write creates a valid JSON file with correct data and snake_case keys."""
    file_path = tmp_path / "output_json"
    json_writer = JSONOutput()
    json_writer.write(str(file_path), sample_full_analysis, analysis_metadata)
    output_file = file_path.with_suffix(".json")
    assert output_file.exists()
    with open(output_file, "r") as f:
        data = json.load(f)
    # Check summary keys are snake_case
    summary_keys = set(data["summary"].keys())
    expected_keys = {
        "file",
        "avg_centroid_distance",
        "totmgterritoryvol",
        "totunoccupiedvol",
        "percentoccupiedvol",
    }
    assert expected_keys.issubset(summary_keys)
    # Check cell keys are snake_case
    cell_keys = set(data["cells"][0].keys())
    expected_cell_keys = {
        "cellterritoryvol",
        "cellvolumes",
        "ramificationindex",
        "numofendpoints",
        "numofbranchpoints",
        "avgbranchlength",
        "maxbranchlength",
        "minbranchlength",
    }
    assert expected_cell_keys.issubset(cell_keys)
    # Check values
    assert data["summary"]["file"] == sample_full_analysis.summary.file
    assert (
        data["cells"][1]["cellterritoryvol"]
        == sample_full_analysis.cells[1].cell_territory_volume
    )


def test_json_output_write_custom_config(tmp_path, sample_full_analysis, analysis_metadata):
    """Test that JSONOutput.write with custom config writes correct snake_case keys."""
    summary_config = AnalysisSummaryConfig(
        file_txt="File",
        avg_centroid_distance_txt="Average Distance",
        total_territorial_volume_txt="Total Territory",
        total_unoccupied_volume_txt="Unoccupied",
        percent_occupied_volume_txt="Percentage Occupied",
    )
    per_cell_config = CellAnalysisConfig(
        cell_territory_volume_txt="Cell Territory",
        cell_volume_txt="Cell Volume",
        ramification_index_txt="Ramification",
        number_of_endpoints_txt="Endpoints",
        number_of_branches_txt="Branches",
        avg_branch_length_txt="Average Length",
        max_branch_length_txt="Maximum Length",
        min_branch_length_txt="Minimum Length",
    )
    file_path = tmp_path / "output_json_custom"
    json_writer = JSONOutput(
        summary_config=summary_config,
        per_cell_config=per_cell_config,
        indent=2,
    )
    json_writer.write(str(file_path), sample_full_analysis, analysis_metadata)
    output_file = file_path.with_suffix(".json")
    with open(output_file, "r") as f:
        data = json.load(f)

    assert "file" in data["summary"]
    assert "average_distance" in data["summary"]
    assert "total_territory" in data["summary"]
    assert "unoccupied" in data["summary"]
    assert "percentage_occupied" in data["summary"]
    assert "cell_territory" in data["cells"][0]
    assert "cell_volume" in data["cells"][0]
    assert "ramification" in data["cells"][0]
    assert "endpoints" in data["cells"][0]
    assert "branches" in data["cells"][0]
    assert "average_length" in data["cells"][0]
    assert "maximum_length" in data["cells"][0]
    assert "minimum_length" in data["cells"][0]


@pytest.mark.parametrize(
    "path, expected",
    [("file.xlsx", "file.xlsx"), ("file", "file.xlsx"), ("file.txt", "file.txt.xlsx")],
)
def test_excel_output_create_path(path, expected):
    """Test ExcelOutput._create_path returns correct file paths with .xlsx extension."""
    excel_writer = ExcelOutput()
    assert excel_writer._create_path(path) == expected


@pytest.mark.parametrize(
    "input_txt,expected",
    [
        ("File", "file"),
        ("Average Distance", "average_distance"),
        ("Total Territory", "total_territory"),
        ("Unoccupied", "unoccupied"),
        ("Percentage Occupied", "percentage_occupied"),
        ("Cell Territory", "cell_territory"),
        ("Cell Volume", "cell_volume"),
        ("Ramification", "ramification"),
        ("Maximum Length", "maximum_length"),
        ("Minimum   Length", "minimum_length"),
        ("  Extra   Spaces  ", "extra_spaces"),
        ("áéíóúüñ", "aeiouun"),
        ("Header-With--Symbols!", "headerwithsymbols"),
    ],
)
def test_json_output_to_snake_case(input_txt, expected):
    """Test JSONOutput._to_snake_case converts strings to lowercase snake_case and removes accents."""
    json_writer = JSONOutput()
    assert json_writer._to_snake_case(input_txt) == expected


def test_branch_lengths_xlsx_output_write(tmp_path, sample_full_analysis, analysis_metadata):
    """Test that BranchLengthsXlsxOutput writes correct branch lengths."""
    file_path = tmp_path / "branch_lengths_test"
    writer = BranchLengthsXlsxOutput()
    writer.write(str(file_path), sample_full_analysis, analysis_metadata)

    # Check file exists
    output_file = file_path.parent / (file_path.name + "_branch_lengths.xlsx")
    assert output_file.exists()

    # Check contents
    wb = load_workbook(output_file)
    ws = wb.active

    # Should have 2 rows (one per cell)
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2

    # Cell 0 has 3 branches - filter out None values
    row0 = [cell for cell in rows[0] if cell is not None]
    assert len(row0) == 3
    assert row0[0] == 8.5
    assert row0[1] == 12.3
    assert row0[2] == 9.2

    # Cell 1 has 4 branches - filter out None values
    row1 = [cell for cell in rows[1] if cell is not None]
    assert len(row1) == 4
    assert row1[0] == 10.1
    assert row1[1] == 11.5
    assert row1[2] == 9.8
    assert row1[3] == 12.6

def test_branch_lengths_xlsx_output_create_path():
    """Test BranchLengthsXlsxOutput._create_path adds _branch_lengths.xlsx suffix."""
    writer = BranchLengthsXlsxOutput()
    assert writer._create_path("file") == "file_branch_lengths.xlsx"
    assert writer._create_path("file.xlsx") == "file_branch_lengths.xlsx"
    assert writer._create_path("file.txt") == "file.txt_branch_lengths.xlsx"


def test_branch_lengths_xlsx_output_get_name():
    """Test BranchLengthsXlsxOutput.get_name returns correct name."""
    assert BranchLengthsXlsxOutput.get_name() == "Branch Lengths XLSX"


def test_branch_lengths_xlsx_output_auto_discovery():
    """Test that the new writer is automatically discovered."""
    writers = OutputWriter.get_writers()
    writer_names = [w.get_name() for w in writers]
    assert "Branch Lengths XLSX" in writer_names
